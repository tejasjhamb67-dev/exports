# -*- coding: utf-8 -*-
"""Build Deliverable 2: Stock_Analysis_Template.xlsx
A sophisticated, reusable single-stock template. Structure follows the CFA Institute
'Equity Research Report Essentials' sections, biased toward markets / flows / deals / country.
Tabs: How to Use · Analysis · Financials · Peer Comps.
Blue = your hardcoded inputs. Yellow = key assumptions. Black = formulas. Grey italic = prompts.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; TEAL="22505A"; PURPLE="4A3F6B"; BAND="F2F5F8"
INPUT="FFF7DE"; BLUE="0000FF"; GREEN="008000"; GREYTX="7A7A7A"; FONT="Arial"

def F(sz=10,b=False,i=False,c="222222"): return Font(name=FONT,size=sz,bold=b,italic=i,color=c)
hdr=F(10,True,c="FFFFFF"); title=F(20,True,c=NAVY); sub=F(10,True,i=True,c=GREYTX)
body=F(10); body_b=F(10,True,c="111111"); prompt=F(9,i=True,c=GREYTX); blue=F(10,c=BLUE)
sec=F(11,True,c="FFFFFF")

navy=PatternFill("solid",fgColor=NAVY); teal=PatternFill("solid",fgColor=TEAL)
purp=PatternFill("solid",fgColor=PURPLE); band=PatternFill("solid",fgColor=BAND)
inp=PatternFill("solid",fgColor=INPUT)
thin=Side(style="thin",color="D0D7DE"); box=Border(thin,thin,thin,thin)
topL=Alignment("left","top",wrap_text=True); midL=Alignment("left","center",wrap_text=True)
midC=Alignment("center","center",wrap_text=True); topC=Alignment("center","top",wrap_text=True)

wb=Workbook()

# ---- helpers ----
def bar(ws,row,text,span=8,fill=navy,f=sec,start=2):
    a=get_column_letter(start); b=get_column_letter(start+span-1)
    ws.merge_cells(f"{a}{row}:{b}{row}")
    c=ws[f"{a}{row}"]; c.value=text; c.font=f; c.fill=fill; c.alignment=Alignment("left","center")
    ws.row_dimensions[row].height=22
def label(ws,row,col,text,f=body_b):
    c=ws.cell(row,col); c.value=text; c.font=f; c.alignment=midL; return c
def infield(ws,row,col,span=1,numfmt=None,align=midC):
    a=get_column_letter(col); b=get_column_letter(col+span-1)
    if span>1: ws.merge_cells(f"{a}{row}:{b}{row}")
    c=ws.cell(row,col); c.fill=inp; c.border=box; c.font=blue; c.alignment=align
    if numfmt: c.number_format=numfmt
    return c
def textarea(ws,row,col,span,height,promptText):
    a=get_column_letter(col); b=get_column_letter(col+span-1)
    ws.merge_cells(f"{a}{row}:{b}{row}")
    c=ws.cell(row,col); c.fill=inp; c.border=box; c.font=prompt; c.alignment=topL; c.value=promptText
    ws.row_dimensions[row].height=height
    return c

# ============================================================ HOW TO USE
g=wb.active; g.title="How to Use"; g.sheet_view.showGridLines=False
for col,w in {"A":2,"B":24,"C":92}.items(): g.column_dimensions[col].width=w
g["B2"]="Single-Stock Analysis — Template"; g["B2"].font=title
g["B3"]="Reusable equity research one-pager · fundamentals-first, lightly technical"; g["B3"].font=sub
guide=[
 ("Purpose","Drop in any ticker and build a clean, decision-ready analysis. Reusable across every stock — "
            "duplicate the 'Analysis' + 'Financials' tabs per name, or overwrite for a fresh one."),
 ("Structure","Follows the CFA Institute 'Equity Research Report Essentials' sections, tuned to your focus: "
              "Markets · Flows · Deals · Country."),
 ("Colour code","BLUE text = your hardcoded inputs.  YELLOW fill = input / assumption cells.  "
                "BLACK = formulas (don't overwrite).  GREY italic = prompts telling you what to write."),
 ("Auto-calcs","Upside %, Market Cap, growth %, margins %, and valuation multiples are formulas — they update "
               "as you type inputs. Everything else is free-text or a number you supply."),
 ("Analysis tab","The hero one-pager: snapshot & recommendation, business, markets, country/geo, competitive "
                 "positioning (Porter mini), deals, capital & money flows, thesis, valuation, risks, ESG, technicals."),
 ("Financials tab","A light 3-year-history + 3-year-forecast skeleton. Enter the blue cells; growth, margins and "
                   "per-share/valuation lines calculate themselves."),
 ("Peer Comps tab","A small relative-valuation table — list 4-6 peers, enter multiples, get an average to anchor "
                   "your target."),
 ("Workflow","1) Fill the snapshot + recommendation.  2) Write the qualitative sections.  3) Enter financials & "
             "peers.  4) Set valuation + target.  5) Sanity-check the technicals strip.  6) Write the one-line thesis."),
 ("Not advice","A structured worksheet, not investment advice. Verify every figure against primary filings."),
]
r=5
for k,v in guide:
    g[f"B{r}"]=k; g[f"B{r}"].font=body_b; g[f"B{r}"].alignment=topL
    g[f"C{r}"]=v; g[f"C{r}"].font=body; g[f"C{r}"].alignment=topL
    g.row_dimensions[r].height=15+14*(1+len(v)//98); r+=1
# example row
g[f"B{r+1}"]="Example (fill format)"; g[f"B{r+1}"].font=body_b
ex=[("Ticker","NVDA"),("Recommendation","Buy / Hold / Sell"),("Current Price","$X.XX"),
    ("Target Price","$X.XX"),("Rev Growth %","enter 0.25 for 25%"),("Gross Margin %","enter 0.75 for 75%")]
r+=2
for k,v in ex:
    g[f"B{r}"]=k; g[f"B{r}"].font=body; g[f"C{r}"]=v; g[f"C{r}"].font=F(10,i=True,c=BLUE); r+=1

# ============================================================ ANALYSIS
a=wb.create_sheet("Analysis"); a.sheet_view.showGridLines=False
widths={"A":2,"B":22,"C":20,"D":18,"E":18,"F":18,"G":18,"H":20,"I":2}
for col,w in widths.items(): a.column_dimensions[col].width=w
NC=8  # content columns B..I? we use B..H (span 7) mostly; use span=7
SPAN=7
a["B1"]="Stock Analysis"; a["B1"].font=title; a.row_dimensions[1].height=26
a.merge_cells("B1:E1")
# top identity band
label(a,2,2,"Company"); infield(a,2,3,2,align=midL).value=None
label(a,2,5,"Ticker"); infield(a,2,6).value=None
label(a,2,7,"As of"); infield(a,2,8,numfmt="yyyy-mm-dd").value=None
for rr in (2,): pass
label(a,3,2,"Exchange"); infield(a,3,3)
label(a,3,5,"Analyst"); infield(a,3,6,2,align=midL)
# recommendation KPI strip
bar(a,5,"RECOMMENDATION SNAPSHOT",SPAN)
kpi=[("Recommendation","text"),("Conviction (1-5)","0"),("Current Price","$#,##0.00"),
     ("Target Price","$#,##0.00"),("Upside %","0.0%"),("Time Horizon","text")]
c0=2
for i,(k,fmt) in enumerate(kpi):
    col=c0+i
    lc=a.cell(6,col); lc.value=k; lc.font=body_b; lc.alignment=midC; lc.fill=band; lc.border=box
    vc=infield(a,7,col, numfmt=None if fmt in("text","0") else fmt)
    if k=="Upside %":
        vc.value="=IFERROR(E7/D7-1,\"\")"; vc.font=F(10,b=True); vc.number_format="0.0%"
    if fmt=="0": vc.number_format="0"
a.row_dimensions[6].height=28

# SECTION 1 — BASIC INFO
bar(a,9,"1 · BASIC INFORMATION  ·  liquidity · float · ownership",SPAN)
basic=[("Sector",3,None),("Industry",6,None),
       ("Country / HQ",3,None),("Reporting Currency",6,None),
       ("Shares Outstanding (m)",3,'#,##0.0'),("Free Float %",6,'0.0%'),
       ("Avg Daily Volume (m)",3,'#,##0.00'),("Beta",6,'0.00'),
       ("52-Week Low",3,'$#,##0.00'),("52-Week High",6,'$#,##0.00'),
       ("Net Debt / (Cash) (m)",3,'#,##0'),("Fiscal Year End",6,None)]
rr=10; toggles=0
for i,(k,col,fmt) in enumerate(basic):
    if col==3 and i>0: rr+=1
    label(a,rr,2 if col==3 else 5,k)
    infield(a,rr,col,numfmt=fmt)
rr+=1
label(a,rr,2,"Market Cap (m)")
mc=infield(a,rr,3,numfmt='#,##0'); mc.value="=IFERROR(D7*C12,\"\")"; mc.font=F(10)  # price*shares
label(a,rr,5,"Enterprise Value (m)")
ev=infield(a,rr,6,numfmt='#,##0'); ev.value="=IFERROR(C{0}+C15,\"\")".format(rr); ev.font=F(10)  # mktcap+netdebt
rr+=1
label(a,rr,2,"Major Shareholders")
textarea(a,rr,3,6,44,"e.g. Vanguard 8%, BlackRock 7%, founder/insiders X% — note any large float overhang or control block.")
rr+=1
label(a,rr,2,"Liquidity Note")
textarea(a,rr,3,6,40,"Float vs market cap, ADR vs local liquidity, index membership, any thin-float or stress-liquidity caveats.")

# SECTION 2 — BUSINESS
rr+=2; bar(a,rr,"2 · BUSINESS DESCRIPTION  ·  what they do · revenue & cost drivers",SPAN)
rr+=1; textarea(a,rr,2,SPAN,70,"What the company does; its core products/services; the KEY DRIVERS of revenue and of costs; unit economics. Source from filings.")
rr+=1; label(a,rr,2,"Revenue by Segment"); label(a,rr,5,"% of Revenue (enter 0.30 = 30%)",F(9,i=True,c=GREYTX))
for seg in range(4):
    rr+=1; infield(a,rr,2,2,align=midL); infield(a,rr,4,numfmt='0.0%')
    a.cell(rr,2).value=None

# SECTION 3 — MARKETS
rr+=2; bar(a,rr,"3 · MARKETS & DEMAND DRIVERS  ·  TAM · end-markets · secular growth",SPAN)
rr+=1; textarea(a,rr,2,SPAN,66,"End-markets served; total addressable market and growth; the secular demand drivers (e.g. AI capex, electrification); cyclicality; where in the cycle we are.")

# SECTION 4 — COUNTRY / GEO
rr+=2; bar(a,rr,"4 · COUNTRY / GEOGRAPHIC EXPOSURE  ·  revenue by region · geopolitics",SPAN)
rr+=1; label(a,rr,2,"Revenue by Region"); label(a,rr,5,"% (0.40 = 40%)",F(9,i=True,c=GREYTX))
for reg in ["Americas","EMEA","Greater China","Rest of Asia-Pac"]:
    rr+=1; c=a.cell(rr,2); c.value=reg; c.font=body; c.border=box; c.alignment=midL
    a.merge_cells(f"B{rr}:C{rr}"); infield(a,rr,4,numfmt='0.0%')
rr+=1; label(a,rr,2,"Geo / Geopolitical Notes")
textarea(a,rr,3,6,46,"Manufacturing vs sales geography; export-control / tariff / sanction exposure; FX sensitivity; regulatory regime by country.")

# SECTION 5 — INDUSTRY & COMPETITIVE POSITIONING
rr+=2; bar(a,rr,"5 · INDUSTRY & COMPETITIVE POSITIONING  ·  Porter's Five Forces · moat",SPAN)
rr+=1; label(a,rr,2,"Five Forces",body_b); label(a,rr,4,"Intensity 1-5",F(9,i=True,c=GREYTX)); label(a,rr,5,"Note",F(9,i=True,c=GREYTX))
forces=["Threat of new entrants","Supplier power","Buyer power","Threat of substitutes","Competitive rivalry"]
for fce in forces:
    rr+=1; c=a.cell(rr,2); c.value=fce; c.font=body; c.border=box; c.alignment=midL; a.merge_cells(f"B{rr}:C{rr}")
    infield(a,rr,4,numfmt='0')
    textarea(a,rr,5,4,16,"")
rr+=1; label(a,rr,2,"Economic Moat")
textarea(a,rr,3,6,44,"Source of durable advantage (scale, brand, switching costs, IP, network effect, cost leadership) — Buffett's 'moat'. How breachable?")
rr+=1; label(a,rr,2,"Key Competitors & Share")
textarea(a,rr,3,6,34,"Peer set; approximate market shares; share trend (gaining/losing); pricing & capacity dynamics.")

# SECTION 6 — DEALS
rr+=2; bar(a,rr,"6 · DEALS, PARTNERSHIPS & M&A",SPAN)
rr+=1
for i,h in enumerate(["Date","Counterparty","Type","Value","Note / Strategic Rationale"]):
    span=1 if i<4 else 3
    c=a.cell(rr,2+i if i<4 else 6); c.value=h; c.font=hdr; c.fill=teal; c.alignment=midC; c.border=box
    if i==4: a.merge_cells(f"F{rr}:H{rr}")
for k in range(4):
    rr+=1
    infield(a,rr,2,numfmt='yyyy-mm'); infield(a,rr,3,align=midL); infield(a,rr,4,align=midL); infield(a,rr,5,align=midL)
    infield(a,rr,6,3,align=midL)

# SECTION 7 — CAPITAL & MONEY FLOWS
rr+=2; bar(a,rr,"7 · CAPITAL & MONEY FLOWS  ·  capex · buybacks · ownership · fund flows",SPAN)
rr+=1
flowitems=[("Capex trend / intensity",3),("Dividend & buyback policy",3),
           ("Institutional ownership %",3),("Insider ownership %",3),
           ("Short interest %",3),("Recent fund / ETF flow bias",3)]
# two-column label:input then a text note
pairs=[("Capex ($m, trend)",'#,##0'),("Buyback + Dividend ($m)",'#,##0'),
       ("Institutional Own %",'0.0%'),("Insider Own %",'0.0%'),
       ("Short Interest %",'0.0%'),("Days to Cover",'0.0')]
for i,(k,fmt) in enumerate(pairs):
    if i%2==0: rr+=1
    col=2 if i%2==0 else 5
    label(a,rr,col,k); infield(a,rr,col+1,numfmt=fmt)
rr+=1; label(a,rr,2,"Flows / Capital-Allocation Note")
textarea(a,rr,3,6,46,"Capex cycle & funding (organic vs debt vs partners); buyback pace; ownership concentration; notable 13F / ETF / southbound flow; capital-allocation quality.")

# SECTION 8 — INVESTMENT SUMMARY / THESIS
rr+=2; bar(a,rr,"8 · INVESTMENT SUMMARY & THESIS  ·  why mispriced · catalysts",SPAN)
rr+=1; label(a,rr,2,"Core Thesis (1-2 lines)")
textarea(a,rr,3,6,34,"The one-paragraph 'why own it' — the variant view.")
rr+=1; label(a,rr,2,"What the Market Misses")
textarea(a,rr,3,6,40,"What is the market NOT properly discounting, and what will make it re-price?")
rr+=1; label(a,rr,2,"Catalysts (next 6-12m)")
textarea(a,rr,3,6,40,"Dated catalysts: earnings, product ramps, contract wins, capex guides, regulatory/geopolitical events.")

# SECTION 9 — VALUATION
rr+=2; bar(a,rr,"9 · VALUATION  ·  relative multiples + quick DCF",SPAN)
rr+=1
for i,h in enumerate(["Metric","This Co.","Peer Avg","5-Yr Avg","Implied / Comment"]):
    c=a.cell(rr,2+i if i<4 else 6); c.value=h; c.font=hdr; c.fill=teal; c.alignment=midC; c.border=box
    if i==4: a.merge_cells(f"F{rr}:H{rr}")
for m,fmt in [("P/E",'0.0"x"'),("EV/Sales",'0.0"x"'),("EV/EBITDA",'0.0"x"'),
              ("P/FCF",'0.0"x"'),("FCF Yield",'0.0%'),("PEG",'0.00')]:
    rr+=1
    c=a.cell(rr,2); c.value=m; c.font=body; c.border=box; c.alignment=midL
    infield(a,rr,3,numfmt=fmt); infield(a,rr,4,numfmt=fmt); infield(a,rr,5,numfmt=fmt); infield(a,rr,6,3,align=midL)
rr+=1; label(a,rr,2,"Quick DCF")
dcf_start=rr+1
for k,fmt,val in [("WACC",'0.0%',None),("Terminal growth",'0.0%',None),
                  ("FY+1 FCF ($m)",'#,##0',None)]:
    rr+=1; label(a,rr,2,k); infield(a,rr,3,numfmt=fmt)
rr+=1; label(a,rr,2,"Implied EV — perpetuity (m)")
iv=infield(a,rr,3,numfmt='#,##0')
iv.value=f"=IFERROR(C{dcf_start+2}*(1+C{dcf_start+1})/(C{dcf_start}-C{dcf_start+1}),\"\")"; iv.font=F(10)
label(a,rr,5,"(Gordon-growth check vs market cap)",F(9,i=True,c=GREYTX))

# SECTION 10 — RISKS
rr+=2; bar(a,rr,"10 · INVESTMENT RISKS",SPAN)
rr+=1
for i,h in enumerate(["Risk","Type","Severity 1-5","Mitigant / Note"]):
    col={0:2,1:4,2:5,3:6}[i]; c=a.cell(rr,col); c.value=h; c.font=hdr; c.fill=teal; c.alignment=midC; c.border=box
    if i==0: a.merge_cells(f"B{rr}:C{rr}")
    if i==3: a.merge_cells(f"F{rr}:H{rr}")
for k in range(5):
    rr+=1; infield(a,rr,2,2,align=midL); infield(a,rr,4,align=midC); infield(a,rr,5,numfmt='0'); infield(a,rr,6,3,align=midL)

# SECTION 11 — ESG
rr+=2; bar(a,rr,"11 · ENVIRONMENTAL · SOCIAL · GOVERNANCE",SPAN)
rr+=1
for pil in ["Environmental","Social","Governance"]:
    c=a.cell(rr,2); c.value=pil; c.font=body_b; c.border=box; c.alignment=midL; a.merge_cells(f"B{rr}:C{rr}")
    textarea(a,rr,4,5,30,f"Key {pil.lower()} factors, risks and any red flags.")
    rr+=1

# SECTION 12 — TECHNICALS
bar(a,rr,"12 · TECHNICAL STRIP  ·  a light overlay on the fundamentals",SPAN)
rr+=1
tech=[("50-day MA",'$#,##0.00'),("200-day MA",'$#,##0.00'),("RSI(14)",'0'),
      ("Support",'$#,##0.00'),("Resistance",'$#,##0.00'),("Trend (Up/Down/Range)",None)]
for i,(k,fmt) in enumerate(tech):
    if i%3==0: rr+=1
    col=2+(i%3)*2
    label(a,rr,col,k); infield(a,rr,col+1,numfmt=fmt)
rr+=1; label(a,rr,2,"% vs 52-Week Range")
pos=infield(a,rr,3,numfmt='0.0%')
pos.value="=IFERROR((D7-C14)/(F14-C14),\"\")"; pos.font=F(10)  # (price-52wLow C14)/(52wHigh F14 - 52wLow C14)
label(a,rr,5,"0% = at 52w low, 100% = at 52w high",F(9,i=True,c=GREYTX))

a.freeze_panes="A5"

# ============================================================ FINANCIALS
fin=wb.create_sheet("Financials"); fin.sheet_view.showGridLines=False
fin.column_dimensions["A"].width=2; fin.column_dimensions["B"].width=30
for col in "CDEFGH": fin.column_dimensions[col].width=13
fin["B1"]="Financial Summary — history & forecast"; fin["B1"].font=title; fin.row_dimensions[1].height=26
fin["B2"]="Blue = input. Black = formula (growth, margins, per-share, multiples auto-calc)."; fin["B2"].font=sub
bar(fin,4,"Enter fiscal years in row 5, then the blue lines. Everything else calculates.",7,navy,F(10,True,c="FFFFFF"))
years_row=5
fin.cell(years_row,2).value="($ millions unless noted)"; fin.cell(years_row,2).font=body_b
labs=["FY-2 (A)","FY-1 (A)","FY0 (A)","FY+1 (E)","FY+2 (E)","FY+3 (E)"]
for i,lb in enumerate(labs):
    c=fin.cell(years_row,3+i); c.value=lb; c.font=hdr; c.fill=navy; c.alignment=midC; c.border=box
cols=[get_column_letter(3+i) for i in range(6)]

def line(row,label_txt,fmt='#,##0',inputrow=True,formula=None):
    c=fin.cell(row,2); c.value=label_txt; c.font=body_b; c.alignment=midL
    for i,col in enumerate(cols):
        cell=fin.cell(row,3+i); cell.border=box; cell.number_format=fmt
        if formula:
            cell.value=formula(col,i); cell.font=F(10)
        else:
            cell.fill=inp; cell.font=blue
    return row

r=6
r_rev=r; line(r,"Revenue"); r+=1
line(r,"  YoY Growth %",'0.0%',False, lambda col,i:(f'=IFERROR({col}{r_rev}/{cols[i-1]}{r_rev}-1,"")' if i>0 else '')); r+=1
r_gp=r; line(r,"Gross Profit"); r+=1
line(r,"  Gross Margin %",'0.0%',False, lambda col,i:f'=IFERROR({col}{r_gp}/{col}{r_rev},"")'); r+=1
r_oi=r; line(r,"Operating Income"); r+=1
line(r,"  Operating Margin %",'0.0%',False, lambda col,i:f'=IFERROR({col}{r_oi}/{col}{r_rev},"")'); r+=1
r_ni=r; line(r,"Net Income"); r+=1
r_sh=r; line(r,"Diluted Shares (m)"); r+=1
line(r,"  EPS ($)",'$#,##0.00',False, lambda col,i:f'=IFERROR({col}{r_ni}/{col}{r_sh},"")'); r+=1
r_fcf=r; line(r,"Free Cash Flow"); r+=1
line(r,"  FCF Margin %",'0.0%',False, lambda col,i:f'=IFERROR({col}{r_fcf}/{col}{r_rev},"")'); r+=1
r+=1
bar(fin,r,"VALUATION CHECK  ·  enter current price + net debt",7,teal,F(10,True,c="FFFFFF")); r+=1
label(fin,r,2,"Current Price ($)"); pcell=fin.cell(r,3); pcell.fill=inp; pcell.font=blue; pcell.border=box; pcell.number_format='$#,##0.00'; r_price=r; r+=1
label(fin,r,2,"Net Debt / (Cash) (m)"); ncell=fin.cell(r,3); ncell.fill=inp; ncell.font=blue; ncell.border=box; ncell.number_format='#,##0'; r_nd=r; r+=1
line(r,"Market Cap (m)",'#,##0',False, lambda col,i:f'=IFERROR($C${r_price}*{col}{r_sh},"")'); r_mc=r; r+=1
line(r,"Enterprise Value (m)",'#,##0',False, lambda col,i:f'=IFERROR({col}{r_mc}+$C${r_nd},"")'); r_ev=r; r+=1
line(r,"P/E (x)",'0.0"x"',False, lambda col,i:f'=IFERROR({col}{r_mc}/{col}{r_ni},"")'); r+=1
line(r,"EV/Sales (x)",'0.0"x"',False, lambda col,i:f'=IFERROR({col}{r_ev}/{col}{r_rev},"")'); r+=1
line(r,"FCF Yield %",'0.0%',False, lambda col,i:f'=IFERROR({col}{r_fcf}/{col}{r_mc},"")'); r+=1

# ============================================================ PEER COMPS
pc=wb.create_sheet("Peer Comps"); pc.sheet_view.showGridLines=False
pc.column_dimensions["A"].width=2; pc.column_dimensions["B"].width=26
for col in "CDEFGH": pc.column_dimensions[col].width=13
pc["B1"]="Peer Comparison"; pc["B1"].font=title; pc.row_dimensions[1].height=26
pc["B2"]="List 4-6 peers; enter the multiples; the average anchors your target."; pc["B2"].font=sub
heads=["Peer","Mkt Cap (m)","P/E","EV/Sales","EV/EBITDA","Rev Gr %"]
hr=4
for i,h in enumerate(heads):
    c=pc.cell(hr,2+i); c.value=h; c.font=hdr; c.fill=navy; c.alignment=midC; c.border=box
fmts=[None,'#,##0','0.0"x"','0.0"x"','0.0"x"','0.0%']
for k in range(6):
    row=hr+1+k
    for i in range(6):
        c=pc.cell(row,2+i); c.fill=inp; c.border=box; c.font=blue
        if i==0: c.alignment=midL
        else: c.alignment=midC;
        if fmts[i]: c.number_format=fmts[i]
avg_row=hr+7
pc.cell(avg_row,2).value="Peer Average"; pc.cell(avg_row,2).font=body_b; pc.cell(avg_row,2).fill=band; pc.cell(avg_row,2).border=box
for i in range(1,6):
    col=get_column_letter(2+i); c=pc.cell(avg_row,2+i)
    c.value=f'=IFERROR(AVERAGE({col}{hr+1}:{col}{hr+6}),"")'; c.font=F(10,True); c.border=box; c.fill=band
    c.number_format=fmts[i] if fmts[i] else 'General'
pc.cell(avg_row+2,2).value="Tip: apply Peer-Avg P/E × your FY+1 EPS (Financials tab) to sanity-check the target price."
pc.cell(avg_row+2,2).font=F(9,i=True,c=GREYTX)

wb.save("Stock_Analysis_Template.xlsx")
print("saved Stock_Analysis_Template.xlsx  tabs:", wb.sheetnames)
