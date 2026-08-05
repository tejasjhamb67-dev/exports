# -*- coding: utf-8 -*-
"""Build Deliverable 1: AI_Adoption_India_Repository.xlsx
Compiles all agent-gathered JSONL into a filterable literature repository of REAL links.
Tabs: Read Me · All Sources (filterable, hyperlinked) · Summary (counts).
Every URL is a real search-result link gathered by the research agents; nothing is fabricated.
"""
import json, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

NAVY="1F3A5F"; TEAL="22505A"; BAND="F2F5F8"; GREYTX="7A7A7A"; FONT="Arial"
def F(sz=10,b=False,i=False,c="222222"): return Font(name=FONT,size=sz,bold=b,italic=i,color=c)
hdr=F(10,True,c="FFFFFF"); title=F(20,True,c=NAVY); sub=F(10,True,i=True,c=GREYTX)
body=F(10); body_b=F(10,True,c="111111"); link=F(10,c="1155CC",i=False)
navy=PatternFill("solid",fgColor=NAVY); teal=PatternFill("solid",fgColor=TEAL); band=PatternFill("solid",fgColor=BAND)
thin=Side(style="thin",color="D0D7DE"); box=Border(thin,thin,thin,thin)
topL=Alignment("left","top",wrap_text=True); midL=Alignment("left","center",wrap_text=True); midC=Alignment("center","center",wrap_text=True)

DATA_DIR=os.path.join(os.path.dirname(__file__),"repo_data")

def load_records():
    recs=[]; seen=set(); malformed=0
    for fp in sorted(glob.glob(os.path.join(DATA_DIR,"*.jsonl"))):
        with open(fp,encoding="utf-8") as f:
            for line in f:
                line=line.strip()
                if not line: continue
                try: o=json.loads(line)
                except Exception: malformed+=1; continue
                url=(o.get("url") or "").strip()
                if not url or not url.lower().startswith("http"): continue
                key=url.rstrip("/").lower()
                if key in seen: continue
                seen.add(key)
                recs.append({k:(o.get(k) or "").strip() for k in
                    ["sector","industry","company","ticker","title","publisher","date","url","ai_category","summary"]})
    return recs, malformed

recs, malformed = load_records()
recs.sort(key=lambda r:(r["sector"].lower(), r["company"].lower(), r["date"]))
print(f"loaded {len(recs)} unique records ({malformed} malformed lines skipped)")

wb=Workbook()

# ---------------- READ ME ----------------
g=wb.active; g.title="Read Me"; g.sheet_view.showGridLines=False
for col,w in {"A":2,"B":24,"C":96}.items(): g.column_dimensions[col].width=w
g["B2"]="AI Adoption in Listed Indian Companies — Literature Repository"; g["B2"].font=title
g["B3"]="Raw sources, not analysis. India as an ADOPTER of AI across sectors."; g["B3"].font=sub
n_companies=len({r["company"] for r in recs if r["company"]})
n_sectors=len({r["sector"] for r in recs if r["sector"]})
guide=[
 ("What this is","A searchable repository of real, published sources (news, press releases, filings, consulting/industry "
                 "reports) documenting how listed Indian companies use or integrate AI. This is your reading list — do your own analysis."),
 ("How to use","Go to 'All Sources'. Use the filter arrows on the header row to slice by Sector, Company, AI Category, "
               "Publisher or Date. Click any link in the URL column to open the source. 'Summary' has counts by sector, company and AI theme."),
 ("Count", f"{len(recs)} unique sources · {n_companies} companies · {n_sectors} sector groups (this pass)."),
 ("Every link is real","No URL here is fabricated. Each was captured from an actual web-search result by the research agents. "
                        "Some links may sit behind paywalls or move over time — that is the nature of a live literature list."),
 ("Why not 'thousands' yet","This environment caps web search at 200 queries per session (shared across all research agents) "
                            "and blocks direct page-fetching by network policy. So one session has a hard ceiling. To grow this "
                            "toward thousands, re-run the research in FRESH sessions (each gets a new 200-query budget) and append — "
                            "the compiler de-duplicates by URL, so repository just keeps growing."),
 ("To expand","Drop more JSONL files (same 10-key schema) into the 'repo_data' folder and re-run build_repository.py. "
              "New unique links merge in automatically; duplicates are ignored."),
 ("Companion file","India_AI_Adopter_Screener.xlsx — the clean screening template (financials + AI columns + sector map)."),
 ("Not advice","A literature index, not investment advice. Verify claims against primary sources."),
]
r=5
for k,v in guide:
    g[f"B{r}"]=k; g[f"B{r}"].font=body_b; g[f"B{r}"].alignment=topL
    g[f"C{r}"]=v; g[f"C{r}"].font=body; g[f"C{r}"].alignment=topL
    g.row_dimensions[r].height=15+14*(1+len(v)//103); r+=1

# ---------------- ALL SOURCES ----------------
ws=wb.create_sheet("All Sources"); ws.sheet_view.showGridLines=False
cols=[("Sector",16),("Industry",18),("Company",22),("Ticker",10),("AI Category",18),
      ("Title / Headline",52),("Publisher",18),("Date",9),("AI Application (summary)",56),("Link",46)]
ws["A1"]="All Sources"; ws["A1"].font=title
ws["A2"]=f"{len(recs)} real, de-duplicated sources. Filter the header row; click a Link to open."; ws["A2"].font=sub
hrow=3
for i,(h,w) in enumerate(cols,1):
    c=ws.cell(hrow,i); c.value=h; c.font=hdr; c.fill=navy; c.alignment=midC; c.border=box
    ws.column_dimensions[get_column_letter(i)].width=w
keys=["sector","industry","company","ticker","ai_category","title","publisher","date","summary","url"]
for ri,rec in enumerate(recs):
    rr=hrow+1+ri
    for ci,key in enumerate(keys,1):
        c=ws.cell(rr,ci); c.border=box
        if key=="url":
            disp=rec["url"]
            disp=disp if len(disp)<=60 else disp[:57]+"..."
            c.value=disp; c.hyperlink=rec["url"]; c.font=link; c.alignment=midL
        else:
            c.value=rec[key]; c.font=body_b if key=="company" else body
            c.alignment = topL if key in ("title","summary") else (midC if key in ("ticker","date","ai_category") else midL)
        if ri%2:
            if not c.fill or c.fill.fgColor.rgb in (None,"00000000"): c.fill=band
    ws.row_dimensions[rr].height=44
ws.freeze_panes="A4"
ws.auto_filter.ref=f"A{hrow}:{get_column_letter(len(cols))}{hrow+len(recs)}"

# ---------------- SUMMARY ----------------
sm=wb.create_sheet("Summary"); sm.sheet_view.showGridLines=False
sm["B1"]="Summary — coverage counts"; sm["B1"].font=title
def block(title_txt, counter, startcol, topn=None):
    sm.cell(3,startcol).value=title_txt; sm.cell(3,startcol).font=hdr; sm.cell(3,startcol).fill=navy
    sm.cell(3,startcol+1).value="#"; sm.cell(3,startcol+1).font=hdr; sm.cell(3,startcol+1).fill=navy
    sm.cell(3,startcol).alignment=midL; sm.cell(3,startcol+1).alignment=midC
    items=counter.most_common(topn) if topn else sorted(counter.items(), key=lambda x:-x[1])
    for i,(k,v) in enumerate(items):
        rr=4+i
        a=sm.cell(rr,startcol); a.value=k or "(blank)"; a.font=body; a.alignment=midL; a.border=box
        b=sm.cell(rr,startcol+1); b.value=v; b.font=body; b.alignment=midC; b.border=box
    sm.column_dimensions[get_column_letter(startcol)].width=30
    sm.column_dimensions[get_column_letter(startcol+1)].width=6
block("By Sector", Counter(r["sector"] for r in recs), 2)
block("By AI Category", Counter(r["ai_category"] for r in recs), 5)
block("Top 30 Companies", Counter(r["company"] for r in recs), 8, topn=30)
sm.cell(1,2).font=title

wb.save("AI_Adoption_India_Repository.xlsx")
print("saved AI_Adoption_India_Repository.xlsx  tabs:", wb.sheetnames, "| rows:", len(recs))
