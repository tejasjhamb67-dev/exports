# -*- coding: utf-8 -*-
"""Build Deliverable 2: India_AI_Adopter_Screener.xlsx
Clean screening template for listed Indian AI-ADOPTER stocks.
Tabs: Read Me · Screener (blank) · Examples (populated) · Sector AI-Impact Map.
Blue/yellow = your inputs. Black = formulas. Grey italic = prompts.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from india_template_data import EXAMPLES, SECTOR_MAP

NAVY="1F3A5F"; TEAL="22505A"; GREEN_H="1E5631"; BAND="F2F5F8"; INPUT="FFF7DE"
GREYTX="7A7A7A"; FONT="Arial"
def F(sz=10,b=False,i=False,c="222222"): return Font(name=FONT,size=sz,bold=b,italic=i,color=c)
hdr=F(10,True,c="FFFFFF"); title=F(20,True,c=NAVY); sub=F(10,True,i=True,c=GREYTX)
body=F(10); body_b=F(10,True,c="111111"); tick=F(10,True,c=NAVY); prm=F(9,i=True,c=GREYTX)
blue=F(10,c="0000FF")
navy=PatternFill("solid",fgColor=NAVY); teal=PatternFill("solid",fgColor=TEAL)
greenf=PatternFill("solid",fgColor=GREEN_H); band=PatternFill("solid",fgColor=BAND); inp=PatternFill("solid",fgColor=INPUT)
thin=Side(style="thin",color="D0D7DE"); box=Border(thin,thin,thin,thin)
topL=Alignment("left","top",wrap_text=True); midL=Alignment("left","center",wrap_text=True)
midC=Alignment("center","center",wrap_text=True); topC=Alignment("center","top",wrap_text=True)

# score color scale 1..5
SCORE_FILL={5:"1E8449",4:"58D68D",3:"F7DC6F",2:"F0B27A",1:"E59866"}

wb=Workbook()

# ---- column schema ----
IDENT=[("ticker","Ticker",10),("company","Company",26),("sector","Sector",18),("industry","Industry",18)]
FIN=[("CMP (Rs)",'"Rs"#,##0.00',11),("Mkt Cap (Rs Cr)",'#,##0',13),("Revenue TTM (Rs Cr)",'#,##0',14),
     ("Rev CAGR 3Y %",'0.0%',11),("EBITDA Margin %",'0.0%',12),("PAT (Rs Cr)",'#,##0',11),
     ("EPS (Rs)",'"Rs"#,##0.00',10),("P/E",'0.0"x"',8),("P/B",'0.0"x"',8),("ROE %",'0.0%',9),
     ("ROCE %",'0.0%',9),("Debt/Equity",'0.00',10),("Promoter Hold %",'0.0%',12),("Div Yield %",'0.0%',10),
     ("Target (Rs)",'"Rs"#,##0.00',11),("Upside %",'0.0%',9)]
QUAL=[("ai_use","Business Use of AI",46),("ai_category","AI Category",16),("stage","AI Adoption Stage",13),
      ("upside","AI Potential Upside",40),("value_where","Where Value Is Created",34),("lever","Value Lever",16),
      ("impact10y","10-Yr AI Impact on Industry",50),("score","AI Impact (1-5)",9),
      ("conviction","Long Conviction",12),("risk","Key Risk",34),("sources","Sources / Notes",26)]

def sheet_headers(ws,startrow):
    col=2
    order=[]
    for _,name,w in IDENT:
        c=ws.cell(startrow,col); c.value=name; c.font=hdr; c.fill=navy; c.alignment=midC; c.border=box
        ws.column_dimensions[get_column_letter(col)].width=w; order.append(("ident",name)); col+=1
    for name,fmt,w in FIN:
        c=ws.cell(startrow,col); c.value=name; c.font=hdr; c.fill=teal; c.alignment=midC; c.border=box
        ws.column_dimensions[get_column_letter(col)].width=w; order.append(("fin",name,fmt)); col+=1
    for _,name,w in QUAL:
        c=ws.cell(startrow,col); c.value=name; c.font=hdr; c.fill=greenf; c.alignment=midC; c.border=box
        ws.column_dimensions[get_column_letter(col)].width=w; order.append(("qual",name)); col+=1
    ws.row_dimensions[startrow].height=30
    return order

CMP_COL=2+len(IDENT)                 # first FIN col = CMP
TARGET_COL=2+len(IDENT)+len(FIN)-2   # 'Target (Rs)'
UPSIDE_COL=2+len(IDENT)+len(FIN)-1   # 'Upside %'

def fin_cell(ws,row,coloffset,name,fmt,example=False):
    col=2+len(IDENT)+coloffset
    c=ws.cell(row,col); c.border=box; c.alignment=midC; c.fill=inp; c.font=blue; c.number_format=fmt
    if name=="Upside %":
        cmpL=get_column_letter(CMP_COL); tgtL=get_column_letter(TARGET_COL)
        c.value=f'=IFERROR({tgtL}{row}/{cmpL}{row}-1,"")'; c.font=F(10)
    return c

# ============================================================ READ ME
g=wb.active; g.title="Read Me"; g.sheet_view.showGridLines=False
for col,w in {"A":2,"B":24,"C":94}.items(): g.column_dimensions[col].width=w
g["B2"]="India AI-Adopter — Stock Screener"; g["B2"].font=title
g["B3"]="Where to go long on AI as a USER, not a builder · listed Indian companies"; g["B3"].font=sub
guide=[
 ("The lens","India won't own foundation-model AI, but it will be one of the world's biggest ADOPTERS. "
             "This screens listed companies wiring AI into an existing P&L — and which industries that re-rates over ~10 years."),
 ("Tabs","'Screener' = blank template to fill.  'Examples' = real Indian AI-adopters filled in as worked rows.  "
         "'Sector AI-Impact Map' = an industry-level view of where AI matters most and the long ideas under each."),
 ("Colour code","BLUE text / YELLOW fill = your inputs.  BLACK = formulas (Upside % auto-calcs from Target ÷ CMP).  "
                "GREY italic = prompts."),
 ("Financial columns","CMP, Market Cap, Revenue, Rev CAGR, EBITDA margin, PAT, EPS, P/E, P/B, ROE, ROCE, Debt/Equity, "
                      "Promoter holding, Dividend yield, Target, Upside %. Amounts in Rs crore; % as fractions (type 0.18 for 18%)."),
 ("AI columns","Business Use of AI · AI Category · Adoption Stage (Exploring/Piloting/Scaling/Core) · AI Potential Upside "
               "· Where Value Is Created · Value Lever (Cost-out / Revenue / Margin / Product / Moat) · 10-Yr Industry Impact "
               "· AI Impact score (1-5) · Long Conviction · Key Risk · Sources."),
 ("How to use","1) Pick names (use Examples + Sector Map as a start).  2) Fill financials from your data source.  "
               "3) Write the AI columns.  4) Rank by AI Impact × Conviction × valuation to build the long book."),
 ("Companion file","AI_Adoption_India_Repository.xlsx — the raw literature: thousands of links on AI adoption across sectors."),
 ("Not advice","A structured screen, not investment advice. The Examples/Sector views are a starting frame — verify every claim and figure."),
]
r=5
for k,v in guide:
    g[f"B{r}"]=k; g[f"B{r}"].font=body_b; g[f"B{r}"].alignment=topL
    g[f"C{r}"]=v; g[f"C{r}"].font=body; g[f"C{r}"].alignment=topL
    g.row_dimensions[r].height=15+14*(1+len(v)//100); r+=1

# ============================================================ SCREENER (blank + 1 example row)
sc=wb.create_sheet("Screener"); sc.sheet_view.showGridLines=False
sc["B1"]="Screener"; sc["B1"].font=title; sc.row_dimensions[1].height=26
sc["B2"]="Fill the yellow cells. Row 4 is a format example — overwrite or delete it."; sc["B2"].font=sub
order=sheet_headers(sc,3)
# example/placeholder row 4 (grey italic prompts)
er=4
placeholders=["TICKER","Company name","Sector","Industry"]
for i,ph in enumerate(placeholders):
    c=sc.cell(er,2+i); c.value=ph; c.font=prm; c.border=box; c.alignment=midL
for j,(name,fmt,w) in enumerate(FIN):
    fin_cell(sc,er,j,name,fmt)
qcol=2+len(IDENT)+len(FIN)
qprompts=["what AI does in the business","tag","Exploring/Piloting/Scaling/Core","the upside if it works",
          "cost-out / revenue / margin","lever","how AI reshapes the industry by ~2035","1-5","High/Med/Low","main risk","link"]
for j,qp in enumerate(qprompts):
    c=sc.cell(er,qcol+j); c.value=qp; c.font=prm; c.border=box; c.alignment=topL
sc.row_dimensions[er].height=40
# a block of blank rows to fill
for rr in range(5,45):
    for i in range(len(IDENT)):
        c=sc.cell(rr,2+i); c.border=box; c.fill=inp; c.font=blue; c.alignment=midL
    for j,(name,fmt,w) in enumerate(FIN):
        fin_cell(sc,rr,j,name,fmt)
    for j in range(len(QUAL)):
        c=sc.cell(rr,qcol+j); c.border=box; c.fill=inp; c.font=blue; c.alignment=topL
    sc.row_dimensions[rr].height=42
sc.freeze_panes="C4"
sc.auto_filter.ref=f"B3:{get_column_letter(1+len(order))}{44}"

# ============================================================ EXAMPLES (populated)
ex=wb.create_sheet("Examples"); ex.sheet_view.showGridLines=False
ex["B1"]="Examples — real Indian AI adopters"; ex["B1"].font=title; ex.row_dimensions[1].height=26
ex["B2"]="Worked rows (qualitative). Financials left blank for you. A starting frame — verify before acting."; ex["B2"].font=sub
sheet_headers(ex,3)
qcol=2+len(IDENT)+len(FIN)
keys_ident=["ticker","company","sector","industry"]
keys_qual=["ai_use","ai_category","stage","upside","value_where","lever","impact10y","score","conviction","risk"]
for ri,rowd in enumerate(EXAMPLES):
    (tk,co,se,ind,ai_use,cat,stage,ups,vw,lev,imp,score,conv,risk)=rowd
    rr=4+ri
    vals=[tk,co,se,ind]
    for i,v in enumerate(vals):
        c=ex.cell(rr,2+i); c.value=v; c.border=box; c.alignment=midL
        c.font=tick if i==0 else body
    for j,(name,fmt,w) in enumerate(FIN):
        fin_cell(ex,rr,j,name,fmt)   # blank yellow inputs
    qvals=[ai_use,cat,stage,ups,vw,lev,imp,score,conv,risk]
    for j,v in enumerate(qvals):
        c=ex.cell(rr,qcol+j); c.value=v; c.border=box; c.alignment=(midC if j in (2,7,8) else topL); c.font=body
        if j==7:  # AI Impact score color
            c.fill=PatternFill("solid",fgColor=SCORE_FILL.get(int(v),"FFFFFF")); c.alignment=midC; c.font=F(10,True)
    # Sources column left blank yellow
    sc_c=ex.cell(rr,qcol+len(qvals)); sc_c.border=box; sc_c.fill=inp; sc_c.font=blue; sc_c.alignment=topL
    ex.row_dimensions[rr].height=96
ex.freeze_panes="C4"
ex.auto_filter.ref=f"B3:{get_column_letter(1+len(IDENT)+len(FIN)+len(QUAL))}{3+len(EXAMPLES)}"

# ============================================================ SECTOR AI-IMPACT MAP
sm=wb.create_sheet("Sector AI-Impact Map"); sm.sheet_view.showGridLines=False
sm["B1"]="Sector AI-Impact Map — where to hunt for longs"; sm["B1"].font=title; sm.row_dimensions[1].height=26
sm["B2"]="Analyst frame: how much AI reshapes each industry over ~10 years, the value lever, and representative long ideas."; sm["B2"].font=sub
heads=[("Sector",22),("AI Impact (1-5)",12),("Primary Value Lever",20),("Adoption Maturity",16),
       ("10-Year Thesis",70),("Representative Long Ideas (tickers)",42)]
hr=3
for i,(h,w) in enumerate(heads):
    c=sm.cell(hr,2+i); c.value=h; c.font=hdr; c.fill=navy; c.alignment=midC; c.border=box
    sm.column_dimensions[get_column_letter(2+i)].width=w
sm.row_dimensions[hr].height=28
for ri,(sec,score,lever,mat,thesis,ideas) in enumerate(sorted(SECTOR_MAP,key=lambda x:-x[1])):
    rr=hr+1+ri
    cells=[sec,score,lever,mat,thesis,ideas]
    for i,v in enumerate(cells):
        c=sm.cell(rr,2+i); c.value=v; c.border=box
        c.alignment = topL if i in (4,5) else (midC if i in (1,3) else midL)
        c.font=body_b if i==0 else body
        if i==1:
            c.fill=PatternFill("solid",fgColor=SCORE_FILL.get(int(v),"FFFFFF")); c.font=F(10,True); c.alignment=midC
    sm.row_dimensions[rr].height=84
sm.freeze_panes="C4"

wb.save("India_AI_Adopter_Screener.xlsx")
print("saved India_AI_Adopter_Screener.xlsx  tabs:", wb.sheetnames, "| examples:", len(EXAMPLES), "| sectors:", len(SECTOR_MAP))
