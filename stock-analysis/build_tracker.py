# -*- coding: utf-8 -*-
"""Build Deliverable 1: AI_Sector_Stock_Tracker.xlsx
Master glance tab + one tab per sub-sector (as per the source sheet's Column A).
Qualitative fields populated; quantitative + technical fields left blank (yellow) for the user.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from company_data import COMPANIES

# ---------- palette ----------
NAVY   = "1F3A5F"   # primary headers
TEAL   = "22505A"   # quant block
PURPLE = "4A3F6B"   # technicals block
BAND   = "F2F5F8"   # zebra band
INPUT  = "FFF7DE"   # user-input (blank) cells
GREYTX = "666666"
FONT   = "Arial"

hdr_font   = Font(name=FONT, size=10, bold=True, color="FFFFFF")
title_font = Font(name=FONT, size=20, bold=True, color=NAVY)
sub_font   = Font(name=FONT, size=10, italic=True, color=GREYTX)
body       = Font(name=FONT, size=10, color="222222")
body_b     = Font(name=FONT, size=10, bold=True, color="111111")
tick_font  = Font(name=FONT, size=10, bold=True, color=NAVY)

navy_fill  = PatternFill("solid", fgColor=NAVY)
teal_fill  = PatternFill("solid", fgColor=TEAL)
purp_fill  = PatternFill("solid", fgColor=PURPLE)
band_fill  = PatternFill("solid", fgColor=BAND)
input_fill = PatternFill("solid", fgColor=INPUT)

thin = Side(style="thin", color="D0D7DE")
med  = Side(style="thin", color="9AA5B1")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

topL   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
topC   = Alignment(horizontal="center", vertical="top",    wrap_text=True)
midL   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
midC   = Alignment(horizontal="center", vertical="center", wrap_text=True)

SECTORS = ["Hardware","Infrastructure","Networking","Software","Cybersecurity",
           "Industrial AI","Mobility AI","Healthcare AI","China AI"]

wb = Workbook()

# ============================================================= COVER
cov = wb.active
cov.title = "Read Me"
cov.sheet_view.showGridLines = False
for c,w in {"A":2,"B":22,"C":90}.items(): cov.column_dimensions[c].width = w
cov["B2"] = "AI & Ancillary Sectors — Stock Tracker"; cov["B2"].font = title_font
cov["B3"] = "Fundamentals-first watchlist · 60 companies · 9 sub-sectors"; cov["B3"].font = sub_font

rows = [
 ("", ""),
 ("What this is", "A replicable, at-a-glance tracker for the AI value chain. Qualitative fields are pre-filled; "
                  "quantitative + technical fields are left blank (highlighted yellow) for you to populate from your data source."),
 ("Universe", "60 unique companies. The source watchlist had 64 rows but 4 were duplicates "
              "(Symbotic, Tempus AI, IQVIA, Recursion) — removed for a clean master list."),
 ("Tabs", "MASTER = every name on one line. Then one tab per sub-sector, as per your sheet's grouping."),
 ("Priorities", "Every qualitative field is biased toward your focus: MARKETS · FLOWS · DEALS · COUNTRY."),
 ("Yellow cells", "Yellow = your inputs. Fill Price, Market Cap, Revenue, margins, multiples, ratings, targets and "
                  "the technical levels. 'Upside %' auto-calculates once you enter Price and Target."),
 ("Qualitative fields", "Business · AI Role · Markets & Demand · Country/Geo · Deals & Partnerships · "
                        "Capital & Money Flows · Moat & Competitors · Risks · Catalysts · Thesis · ESG."),
 ("Quant fields (blank)", "Price · Market Cap · Revenue (TTM) · Rev Growth % · Gross Margin % · EV/Sales · "
                          "P/E · FCF Yield · Rating · Target · Upside %."),
 ("Technical fields (blank)", "52W Low/High · 50-day MA · 200-day MA · RSI(14) · Trend."),
 ("Analysis frame", "Structure follows the CFA Institute 'Equity Research Report Essentials' sections: Basic Info, "
                    "Business, Industry & Competitive Positioning, Investment Summary, Valuation, Financials, Risks, ESG."),
 ("Companion file", "Stock_Analysis_Template.xlsx — a deep single-stock template for full write-ups on any ticker."),
 ("Data caveat", "Qualitative content is analyst-authored (knowledge as of early 2026) as a starting frame — "
                 "verify against primary filings before acting. Nothing here is investment advice."),
]
r = 5
for k,v in rows:
    cov[f"B{r}"] = k; cov[f"B{r}"].font = body_b; cov[f"B{r}"].alignment = topL
    cov[f"C{r}"] = v; cov[f"C{r}"].font = body;   cov[f"C{r}"].alignment = topL
    cov.row_dimensions[r].height = 15 + 14*(1+len(v)//95)
    r += 1
cov["B2"].alignment = Alignment(vertical="center")

# ============================================================= MASTER
QUANT_M = ["Price","Market Cap","P/E","EV/Sales","Rev Growth %","Rating","Target","Upside %"]
ms = wb.create_sheet("MASTER")
ms.sheet_view.showGridLines = False
mh1 = ["Sector","Sub-Industry","Company","Ticker","Exchange","Country","AI Value-Chain Role","Investment Thesis (one-line)"]
headers = mh1 + QUANT_M
ms.append([])  # title row
ms["A1"] = "MASTER — 60 AI & Ancillary Names at a Glance"; ms["A1"].font = title_font
ms["A2"] = "Yellow = your inputs. Upside % auto-calcs from Target ÷ Price."; ms["A2"].font = sub_font
ms.append([])
ms.append(headers)
hrow = 3
for i,h in enumerate(headers,1):
    cell = ms.cell(hrow,i); cell.value=h; cell.font=hdr_font; cell.alignment=midC; cell.border=box
    cell.fill = teal_fill if h in QUANT_M else navy_fill
for i,comp in enumerate(COMPANIES):
    rr = hrow+1+i
    vals = [comp["sector"],comp["subind"],comp["name"],comp["ticker"],comp["exch"],comp["country"],comp["tag"],comp["thesis"]]
    for ci,v in enumerate(vals,1):
        c = ms.cell(rr,ci); c.value=v; c.border=box; c.alignment = midL if ci in (7,8) else midC
        c.font = tick_font if ci in (3,4) else body
        if i%2: c.fill = band_fill
    for qi,q in enumerate(QUANT_M):
        ci = len(mh1)+1+qi
        c = ms.cell(rr,ci); c.border=box; c.alignment=midC; c.fill=input_fill; c.font=body
        if q=="Upside %":
            pcol=get_column_letter(len(mh1)+1); tcol=get_column_letter(len(mh1)+1+QUANT_M.index("Target"))
            c.value=f'=IFERROR({tcol}{rr}/{pcol}{rr}-1,"")'; c.number_format="0.0%"
        elif q in ("Price","Target"): c.number_format='$#,##0.00'
        elif q=="Market Cap": c.number_format='$#,##0,,"B"'
        elif q in ("P/E","EV/Sales"): c.number_format='0.0"x"'
        elif q=="Rev Growth %": c.number_format='0.0%'
mwidths=[13,16,22,9,13,12,26,50,11,13,8,10,12,9,10,10]
for i,w in enumerate(mwidths,1): ms.column_dimensions[get_column_letter(i)].width=w
ms.freeze_panes="D4"
ms.auto_filter.ref=f"A{hrow}:{get_column_letter(len(headers))}{hrow+len(COMPANIES)}"

# ============================================================= SECTOR TABS
QUAL = [("business","Business"),("ai_role","AI Role"),("markets","Markets & Demand Drivers"),
        ("geo","Country / Geo Exposure"),("deals","Key Deals & Partnerships"),
        ("flows","Capital & Money Flows"),("moat","Moat & Key Competitors"),
        ("risks","Key Risks"),("catalyst","Catalysts to Watch"),("thesis","Investment Thesis"),
        ("esg","ESG Note")]
QUANT = ["Price","Market Cap","Revenue (TTM)","Rev Growth %","Gross Margin %","EV/Sales","P/E","FCF Yield %","Rating","Target","Upside %"]
TECH  = ["52W Low","52W High","50-day MA","200-day MA","RSI(14)","Trend"]
IDCOLS = ["Company","Ticker","Exch","Country"]

def build_sector(name):
    data=[c for c in COMPANIES if c["sector"]==name]
    ws=wb.create_sheet(name[:31]); ws.sheet_view.showGridLines=False
    ws["A1"]=f"{name} — {len(data)} name{'s' if len(data)!=1 else ''}"; ws["A1"].font=title_font
    ws["A2"]="Populated: qualitative (Business→ESG).  Blank yellow: your quant + technical inputs.  Upside % auto-calcs."; ws["A2"].font=sub_font
    headers = IDCOLS + [h for _,h in QUAL] + QUANT + TECH
    hrow=3
    for i,h in enumerate(headers,1):
        c=ws.cell(hrow,i); c.value=h; c.font=hdr_font; c.alignment=midC; c.border=box
        if h in QUANT: c.fill=teal_fill
        elif h in TECH: c.fill=purp_fill
        else: c.fill=navy_fill
    nID=len(IDCOLS); nQUAL=len(QUAL)
    for ri,comp in enumerate(data):
        rr=hrow+1+ri
        ids=[comp["name"],comp["ticker"],comp["exch"],comp["country"]]
        for ci,v in enumerate(ids,1):
            c=ws.cell(rr,ci); c.value=v; c.border=box; c.alignment=topL
            c.font = tick_font if ci in (1,2) else body
        for qi,(key,_) in enumerate(QUAL):
            c=ws.cell(rr,nID+1+qi); c.value=comp[key]; c.border=box; c.alignment=topL; c.font=body
        qstart=nID+nQUAL
        for qi,q in enumerate(QUANT):
            ci=qstart+1+qi; c=ws.cell(rr,ci); c.border=box; c.alignment=topC; c.fill=input_fill; c.font=body
            if q=="Upside %":
                pcol=get_column_letter(qstart+1+QUANT.index("Price")); tcol=get_column_letter(qstart+1+QUANT.index("Target"))
                c.value=f'=IFERROR({tcol}{rr}/{pcol}{rr}-1,"")'; c.number_format="0.0%"
            elif q in ("Price","Target"): c.number_format='$#,##0.00'
            elif q=="Market Cap": c.number_format='$#,##0,,"B"'
            elif q=="Revenue (TTM)": c.number_format='$#,##0,,"B"'
            elif q in ("Rev Growth %","Gross Margin %","FCF Yield %"): c.number_format='0.0%'
            elif q in ("EV/Sales","P/E"): c.number_format='0.0"x"'
        tstart=qstart+len(QUANT)
        for ti,t in enumerate(TECH):
            ci=tstart+1+ti; c=ws.cell(rr,ci); c.border=box; c.alignment=topC; c.fill=input_fill; c.font=body
            if t in ("52W Low","52W High","50-day MA","200-day MA"): c.number_format='$#,##0.00'
            elif t=="RSI(14)": c.number_format='0'
        ws.row_dimensions[rr].height=132
    # widths
    idw=[22,9,13,12]
    qualw=[42,30,44,40,44,42,40,38,38,40,34]
    quantw=[11,13,13,11,12,10,8,11,9,10,10]
    techw=[10,10,11,11,9,12]
    ws_widths=idw+qualw+quantw+techw
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="C4"
    ws.row_dimensions[1].height=26

for s in SECTORS: build_sector(s)

wb.save("AI_Sector_Stock_Tracker.xlsx")
print("saved AI_Sector_Stock_Tracker.xlsx  tabs:", wb.sheetnames)
